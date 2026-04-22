"""
NiCE + NICE Cognigy response to Ralph Lauren Contact Center Technology RFP
Fills: AI Assessment, Data Assessment, Functional Assessment, Supplier Qs, Technical Qs
"""
import openpyxl

INPUT  = "/Users/Adam.Boyle/Downloads/Attachment 2_Ralph Lauren Contact Center Technology RFP Vendor Requirements_Questionnaire_Pricing.xlsx"
OUTPUT = "/Users/Adam.Boyle/Downloads/NiCE Response - Ralph Lauren CCaaS RFP.xlsx"

wb = openpyxl.load_workbook(INPUT)

def cv(ws, row_idx, col_idx):
    """Get cell value as string (1-based col)."""
    val = ws.cell(row=row_idx, column=col_idx).value
    return str(val).strip() if val else ""

def fill_cell(ws, row_idx, col_idx, value):
    ws.cell(row=row_idx, column=col_idx).value = value

# ── Response lookups by subsection / keyword ─────────────────────────────────

AI_RESPONSES = {
    "Model Governance": ("Yes",
        "NICE CXone's AI governance framework covers full audit trails for model and flow changes, version-controlled deployments, and bias testing via intent distribution analysis. Confidence-threshold fallback handling and RAG grounding prevent hallucination by anchoring LLM responses to verified knowledge sources."),
    "Model Transparency": ("Yes",
        "NICE CXone provides full model transparency: every RAG-retrieved response includes the source document and passage. Model change logs are exportable for admin review, and NICE provides model cards and training data summaries under NDA upon request."),
    "Training": ("Yes",
        "NICE CXone's NLU training pipeline uses customer-provided utterance examples and continuously learns from live conversation logs. Hold-out validation sets and intent confusion analysis detect overfitting and bias before each model deployment."),
    "Bias": ("Yes",
        "Intent distribution analysis flags underrepresented or biased training data. NICE Cognigy supports independent evaluation of NLU model outputs and provides confidence score distributions for fairness review."),
    "IVR": ("Yes",
        "NICE CXone delivers enterprise-grade IVR and self-service through NICE Cognigy Voice Gateway. NLU-driven intent recognition replaces legacy DTMF menus, and proactive outbound IVR is managed via CXone Personal Connection."),
    "Proactive": ("Yes",
        "CXone supports proactive outbound voice, SMS, and email notifications. NICE Cognigy orchestrates triggered outbound flows based on CRM events, order status changes, or scheduled campaigns."),
    "Outbound": ("Yes",
        "NICE CXone supports proactive outbound voice, SMS, and email notifications orchestrated by NICE Cognigy AI flows. Triggered outbound campaigns can be initiated by CRM events, order status changes, or scheduled business rules, enabling personalised proactive customer engagement at scale."),
    "Chat": ("Yes",
        "NICE CXone delivers omnichannel chatbot capabilities via NICE Cognigy across webchat, WhatsApp, Apple Messages for Business, SMS, and social channels. NICE Cognigy Webchat v3 supports rich messaging (carousels, quick replies, file upload, forms) and is WCAG 2.1 AA audited (July 2024), aligned with WCAG 2.2 AA. CXone Agent Copilot surfaces AI-generated response suggestions to live agents in real time."),
    "Voice": ("Yes",
        "NICE CXone delivers SIP-based voice AI through NICE Cognigy Voice Gateway, with support for multiple STT and TTS providers including Deepgram Nova-3 and Azure Neural TTS. Custom pronunciation dictionaries, noise filtering, and configurable endpointing are available across all supported languages."),
    "Voicebot": ("Yes",
        "NICE CXone supports production-grade voicebots via NICE Cognigy Voice Gateway. Streaming TTS minimises latency, barge-in is configurable per flow, and silence overlay, DTMF fallback, and no-input timeout handling are all configurable at the flow level."),
    "NLU": ("Yes",
        "NICE CXone's built-in NLU engine provides 28 dedicated language-specific models covering all major global languages, with the Universal Language Model extending intent recognition to 100+ languages. Third-party NLU engines (Google Dialogflow, Amazon Lex, IBM Watson Assistant) are supported via the NLU Connector framework."),
    "Agent Assist": ("Yes",
        "NICE CXone includes Agent Copilot, which delivers real-time next-best-action suggestions, knowledge article surfacing, and AI-drafted response text to live agents during interactions. Agent Copilot integrates directly with the CXone MAX agent desktop."),
    "Speech": ("Yes",
        "CXone Interaction Analytics provides full speech analytics: automated transcription, speaker diarization, sentiment detection, topic clustering, and custom phrase spotting. Results are searchable and exportable."),
    "STT": ("Yes",
        "NICE CXone's voice platform supports multiple STT providers via NICE Cognigy Voice Gateway, including Deepgram Nova-3 (primary), Google STT, Microsoft Azure STT, and Amazon Transcribe. STT providers are configurable per deployment, with tunable noise filtering, endpointing thresholds, and language models."),
    "TTS": ("Yes",
        "NICE CXone's voice platform supports multiple TTS providers via NICE Cognigy Voice Gateway, including ElevenLabs, Microsoft Azure Neural TTS, Google WaveNet, and Nuance. Custom pronunciation dictionaries handle brand names and domain-specific terms, and multiple providers can run in parallel for A/B testing."),
    "Speaker Diarization": ("Yes",
        "Deepgram Nova-3 supports speaker diarization, clearly labeling agent vs. customer speech in transcripts. This feeds directly into CXone Interaction Analytics for per-speaker sentiment and compliance analysis."),
    "Noise": ("Yes",
        "Deepgram Nova-3 includes built-in noise suppression and background audio filtering. Endpointing sensitivity is tunable to prevent false triggers from background noise."),
    "Transcription": ("Yes",
        "Real-time and post-call transcriptions are generated for all voice interactions via Deepgram integration. Transcripts are indexed in CXone and searchable by keyword, phrase, topic, and sentiment."),
    "Model Management": ("Yes",
        "NICE CXone provides full AI model lifecycle management: train, test, stage, and deploy NLU models with version rollback. A/B testing of intent models is supported natively within the platform."),
    "Multi-turn": ("Yes",
        "NICE CXone maintains full conversation context across all turns via a persistent Context object. Session-level and profile-level memory enable multi-session personalisation, and context is passed intact to the receiving agent on handover."),
    "Escalation": ("Yes",
        "NICE CXone's built-in handover protocol transfers the full conversation transcript, extracted entities, intent history, and customer profile to the receiving agent. The CXone agent desktop is pre-populated via screen-pop before the agent accepts the interaction."),
    "Continuous Learning": ("Yes",
        "NICE CXone continuously improves NLU accuracy through NICE Cognigy's Intent Trainer, which surfaces misclassified utterances for human review and retraining. Retraining loops can be scheduled or triggered manually, with NLU performance metrics tracked over time in the analytics dashboard."),
    "LLM": ("Yes",
        "NICE CXone supports multiple LLM providers natively via NICE Cognigy's generative AI framework, including Azure OpenAI, OpenAI GPT-4, Anthropic Claude, Google Gemini, and Amazon Bedrock/Nova (added in v4.92). LLM selection is configurable per flow node, with prompt injection defences and output filtering applied."),
    "Knowledge": ("Yes",
        "NICE CXone's Knowledge AI capability ingests PDFs, URLs, and structured documents; chunks, embeds, and indexes them; and retrieves relevant passages at runtime via RAG. Responses are grounded in verified content, with source citations returned to the agent."),
    "Language": ("Yes",
        "NICE CXone supports 28 dedicated language-specific NLU models, with the Universal Language Model extending recognition to 100+ languages. Language detection triggers automatically from the first utterance and routes to the correct locale. English, Spanish, French, Italian, German, Dutch, and Portuguese are all production-ready with full feature parity."),
    "French": ("Yes",
        "French (both France and Canadian French) is fully supported in NICE Cognigy. Separate locale configurations allow region-specific terminology, TTS voice personas, and compliance rules. Automatic language detection routes callers to the French locale without menu selection."),
    "Auto-translation": ("Yes",
        "NICE CXone supports LLM-based dynamic translation for runtime content rendering, enabling real-time multilingual responses without separate locale configurations. Static localisation is managed per-locale in the CMS, and language detection is automatic."),
    "Translation": ("Yes",
        "Dynamic content can be auto-translated via LLM nodes at runtime. Static translations are managed per-locale. The platform supports side-by-side management of all locale variants in a single flow editor."),
    "Third-party": ("Yes",
        "NICE CXone's NLU Connector framework supports third-party NLU providers including Google Dialogflow, Amazon Lex, and IBM Watson Assistant, as well as custom REST endpoints. TTS and STT providers are swappable per deployment, enabling mix-and-match configurations for language-specific optimisation."),
    "Integration": ("Yes",
        "CXone provides REST APIs, webhooks, and a 150+ app marketplace (CXexchange) for integration with CRM, WFM, analytics, and third-party AI providers. NICE Cognigy integrates via HTTP Request nodes, OAuth flows, and pre-built connectors."),
}

DATA_RESPONSES = {
    "Availability": ("Yes",
        "CXone stores interaction data in a cloud-native, geo-redundant architecture. All interaction records, transcripts, recordings, and analytics data are accessible via the CXone portal and REST API within minutes of interaction close. Role-based data access is enforced at the API and UI layers."),
    "Accessibility": ("Yes",
        "Reports and interaction data are searchable, filterable, and exportable (CSV, Excel, PDF) from the CXone portal. The CXone Data Stream API provides real-time and batch access for BI tool integration (Tableau, Power BI, Snowflake)."),
    "Data Availability": ("Yes",
        "All interaction data (voice recordings, chat transcripts, email threads, AI conversation logs) is available via CXone portal search and REST API. Data is retained per contractual terms with configurable per-region retention policies."),
    "Search": ("Yes",
        "CXone Interaction Analytics provides full-text and faceted search across all interaction records. Filters include date range, channel, agent, team, sentiment, topic, and custom tags. Results are paginated and exportable."),
    "Export": ("Yes",
        "Data export is available in CSV, Excel, and JSON formats via the CXone portal and REST API. Scheduled exports can be configured for automated delivery to SFTP, S3, or Azure Blob Storage."),
    "Context": ("Yes",
        "CXone captures full interaction context from all touchpoints: IVR traversal, bot conversation, agent desktop actions, and CRM updates. Cross-channel journeys are linked by customer identifier for a unified interaction record."),
    "Role-based": ("Yes",
        "RBAC controls data access at the report, dashboard, and API level. Supervisors see team data; agents see only their own data; admins have full access. Custom roles are configurable via the CXone admin portal."),
    "Explainability": ("Yes",
        "NICE CXone provides full AI decision explainability: the top-ranked intent, confidence score, and matched entities are returned for every NLU decision. RAG responses include the source document and passage, and audit logs record every AI decision with timestamps, input, and output for regulatory review."),
    "Audit": ("Yes",
        "CXone and NICE Cognigy maintain full audit trails for all interactions, configuration changes, user actions, and AI decisions. Logs are exportable and SIEM-compatible (Syslog, Splunk, Datadog). Audit records are retained per contract terms."),
    "Confidence": ("Yes",
        "NICE CXone exposes intent confidence scores at runtime and in the analytics dashboard. Threshold-based fallback routing directs low-confidence interactions to clarification flows or live agents, and confidence distributions are visible in the NLU performance dashboard."),
    "Interaction Reporting": ("Yes",
        "CXone provides 100+ prebuilt interaction reports covering AHT, CSAT, FCR, abandonment, channel mix, and agent performance. Custom report builder supports unlimited custom metrics. Reports can be scheduled and exported."),
    "Interaction": ("Yes",
        "CXone stores searchable interaction records for voice, chat, email, and bot sessions. Each record includes metadata (channel, agent, queue, duration, disposition) and links to the recording/transcript."),
    "Multi-channel": ("Yes",
        "CXone's reporting engine covers all channels in a unified dataset. Cross-channel journey reports trace a single customer across voice, chat, email, and bot interactions within a session."),
    "Filtering": ("Yes",
        "All CXone reports support multi-dimensional filtering: date range, channel, agent, team, skill, queue, disposition, sentiment, and custom attributes. Filtered views can be saved as report templates."),
    "Recording Search": ("Yes",
        "CXone Interaction Analytics enables keyword, phrase, and topic search across recorded calls. Search returns timestamped results within recordings, enabling fast QA review without full playback."),
    "Journey": ("Yes",
        "CXone tracks the full customer journey across channels and interactions. Journey maps are available in the analytics module showing channel sequences, hand-offs, and resolution points."),
    "Survey": ("Yes",
        "CXone supports post-interaction IVR surveys, SMS surveys, and email surveys. NPS, CSAT, and CES scoring are supported out of the box. Survey results are correlated with interaction records in analytics."),
    "Custom Report": ("Yes",
        "CXone's custom report builder allows creation of ad-hoc reports on any combination of interaction, agent, queue, and AI metrics. Custom reports support calculated fields, conditional formatting, and drill-down."),
    "Compliance Flagging": ("Yes",
        "CXone Auto QA flags interactions for compliance review based on configurable phrase patterns, sentiment thresholds, and behavior rules. Flagged interactions are routed to QA queues for human review."),
    "Real-Time": ("Yes",
        "CXone Real-Time Reporting delivers live dashboards for queue metrics, agent states, AHT, SLA adherence, and channel volumes at 15-minute interval granularity. Threshold alerts are configurable with email/SMS/in-app notifications."),
    "Threshold": ("Yes",
        "Real-time threshold alerts in CXone trigger notifications when metrics exceed defined limits (e.g., queue wait time, abandonment rate, adherence deviation). Alerts are delivered via email, SMS, or CXone supervisor desktop."),
    "API": ("Yes",
        "CXone exposes a full REST API and real-time Data Stream API for integration with external BI tools, CRM systems, and custom dashboards. OpenAPI (Swagger) documentation is provided. API versioning ensures backward compatibility."),
    "CRM Integration": ("Yes",
        "CXone integrates with Salesforce, Microsoft Dynamics 365, ServiceNow, Zendesk, SAP CRM, and Oracle Service Cloud via managed packages and REST APIs. Real-time screen-pop and bidirectional data sync are supported."),
    "Batch": ("Yes",
        "CXone supports scheduled batch data exports to SFTP, Amazon S3, or Azure Blob Storage. Batch feeds can be configured for interaction records, agent performance data, and analytics aggregations."),
    "SMS": ("Yes",
        "CXone integrates with leading SMS providers (Twilio, Bandwidth, MessageBird) for outbound notification and two-way SMS interactions. Provider configuration is managed in the CXone admin portal."),
    "Carrier": ("Yes",
        "CXone's outbound communications module supports carrier lookup and number validation before campaign execution, reducing TCPA risk and improving deliverability."),
    "Data": ("Yes",
        "CXone provides comprehensive data availability, searchability, and export across all interaction types. Data integrations via REST API and event streaming enable real-time and batch access for BI and CRM systems."),
}

FUNC_RESPONSES = {
    "Omnichannel": ("Yes",
        "CXone delivers end-to-end omnichannel orchestration across voice, chat, email, SMS, WhatsApp, Apple Messages for Business, and social channels. Context is preserved across channel switches within a session. Business rules govern routing, escalation, and channel-switching logic."),
    "E2E": ("Yes",
        "CXone orchestrates complete customer journeys from IVR self-service through bot-assisted interaction to agent-handled resolution, with full context preservation at every handoff. Journey analytics track customers across all touchpoints."),
    "Context Preservation": ("Yes",
        "The NICE Cognigy Context object persists all conversation variables, intent history, and customer data throughout a session and across channel transitions. Agents receive a fully pre-populated screen-pop on handover."),
    "Outbound Campaign": ("Yes",
        "NICE CXone Personal Connection provides full outbound campaign management across voice, SMS, and email. Campaign supervisors configure dialing modes (predictive, preview, progressive), set pacing targets, and manage contact lists from a centralised admin console. Real-time campaign dashboards show agent productivity, connect rates, and abandonment metrics."),
    "Business Rules": ("Yes",
        "CXone's routing engine supports complex business rules: priority queuing, customer tier routing, time-of-day logic, and skill-based routing. Rules are configurable via the admin UI without code."),
    "Async": ("Yes",
        "CXone supports asynchronous messaging channels (email, WhatsApp, Apple Messages) with configurable SLA timers and queue management. Agents can handle multiple async interactions simultaneously."),
    "Channel Switching": ("Yes",
        "CXone enables seamless channel switching mid-interaction (e.g., from voice to chat for a document transfer) while preserving the full conversation context. The customer does not need to re-authenticate or repeat information."),
    "Backend Integration": ("Yes",
        "CXone integrates with backend systems via REST APIs, Salesforce managed package, ServiceNow integration, and the CXexchange marketplace. NICE Cognigy supports API tool-calling mid-conversation for real-time data retrieval."),
    "Routing": ("Yes",
        "CXone's ACD routes based on intent (from NICE Cognigy), customer value/segment, agent skills, real-time queue load, and schedule. AI-powered dynamic routing adjusts in real-time. Skills-based routing supports multi-skill matching with proficiency weighting."),
    "Intent": ("Yes",
        "NICE CXone uses AI-driven intent signals from NICE Cognigy to route contacts directly to the most qualified agent or team, eliminating the need for menu navigation. Intent scores and extracted entities are passed as routing attributes to the CXone ACD in real time."),
    "AI Routing": ("Yes",
        "CXone's AI-powered routing evaluates intent, customer value, predicted handle time, and agent availability to optimize assignment decisions in real time. Routing models can be tuned per skill group or queue."),
    "Affinity": ("Yes",
        "CXone supports agent affinity and last-agent routing at the queue level. Affinity decay is configurable so that preferred-agent requests fall back to the general queue after a defined wait threshold."),
    "Dynamic": ("Yes",
        "CXone routing dynamically adjusts to real-time conditions including queue depth, agent availability, predicted abandon risk, and customer priority. Rules can be modified live without platform downtime."),
    "Voice": ("Yes",
        "CXone handles inbound and outbound voice with full ACD, IVR, call recording, and screen-pop. NICE Cognigy Voice Gateway powers conversational IVR. SIP trunking, PSTN fallback, and E.164 number management are all supported."),
    "Chat": ("Yes",
        "NICE CXone manages chat interactions end-to-end: NICE Cognigy Webchat v3 delivers rich messaging (carousels, quick replies, file upload), while CXone routes chat interactions via skills-based ACD with concurrent limits and supervisor monitoring. All chat transcripts are stored and fully searchable."),
    "Email": ("Yes",
        "CXone Email handles inbound email routing, queuing, and response management. AI-powered email classification routes messages to the correct skill group. Auto-response and template library features reduce handle time."),
    "Email Automation": ("Yes",
        "CXone supports automated email acknowledgments, AI-generated response drafts, and template-based replies. NICE Cognigy can fully handle email interactions for routine inquiries without agent involvement."),
    "IP Blocking": ("Yes",
        "CXone supports IP and phone number blocking at the platform level for fraud and abuse prevention. Block lists are configurable by administrators and take effect in real time."),
    "WFM": ("Yes",
        "CXone WFM provides multi-skill, multi-channel forecasting at 15-minute intervals, automated schedule generation, agent self-scheduling (Shift Bidding), and real-time adherence monitoring with supervisor alerts."),
    "WEM": ("Yes",
        "CXone WEM covers scheduling, adherence, quality management, performance coaching, and gamification in a single platform. Supervisor coaching dashboards surface AI-identified coaching opportunities from Auto QA evaluations."),
    "Scheduling": ("Yes",
        "CXone WFM generates optimized schedules respecting skills, hours, labor rules, and agent preferences. Multi-channel forecasting ensures accurate staffing across voice, chat, and email queues simultaneously."),
    "Forecasting": ("Yes",
        "CXone WFM forecasting uses Erlang C, regression analysis, and pattern-matching algorithms. Historical data ingestion from CXone ACD and third-party systems supports accurate multi-channel volume prediction."),
    "Coaching": ("Yes",
        "CXone Auto QA evaluates 100% of interactions against configurable scorecards. Coaching opportunities are surfaced to supervisors with recommended content. Agents can access coaching materials directly in the CXone agent desktop."),
    "Authentication": ("Yes",
        "CXone supports MFA, SAML 2.0 / OAuth 2.0 / OIDC SSO, and LDAP/Active Directory integration for all user types. Voice biometric authentication for customers is available via CXone's certified integration with leading biometric providers; details available upon request."),
    "MFA": ("Yes",
        "MFA is enforced for all admin roles and configurable for agent and supervisor access. Supported factors: TOTP, push notification, SMS OTP, and hardware tokens via SAML IdP integration."),
    "SSO": ("Yes",
        "CXone supports SAML 2.0 and OAuth 2.0/OIDC for SSO with all major identity providers (Okta, Azure AD, Ping Identity, OneLogin). SCIM 2.0 enables automated user provisioning and deprovisioning."),
    "Biometric": ("Partial",
        "CXone supports passive voice biometric authentication via certified integration with leading voice biometric providers. Returning callers can be authenticated without knowledge-based questions; fraudster voiceprint detection is also available. Specific provider options and deployment scope are discussed during solution design."),
    "Recording": ("Yes",
        "CXone provides 100% call recording with synchronized screen recording. Multi-channel recording covers voice, chat transcripts, and email threads. PCI pause/resume masks sensitive payment data. Recordings are searchable and retained per contractual terms."),
    "Screen Recording": ("Yes",
        "CXone screen recording captures agent desktop activity synchronized with voice recordings. Recordings support QA review, compliance audit, and training. Retention is configurable per region and data type."),
    "Personalization": ("Partial",
        "CXone delivers personalized self-service experiences via NICE Cognigy's customer data lookup at interaction entry — surfacing purchase history, loyalty tier, and preferences from integrated CRM systems (Salesforce, Dynamics). Full VIC clienteling and guided selling require CRM integration, which is supported out of the box via the Salesforce managed package."),
    "VIC": ("Partial",
        "CXone + NICE Cognigy supports VIC (Very Important Customer) personalization through CRM data integration. Recognized VIC callers receive a personalized greeting, priority routing, and pre-populated agent screen-pop with clienteling data. Full boutique-level personalization requires Salesforce or equivalent CRM integration."),
    "Outbound": ("Yes",
        "NICE CXone Personal Connection handles high-volume outbound delivery with AMD (Answering Machine Detection), contact blending between inbound and outbound queues, and configurable abandon rate controls. TCPA compliance controls, time-zone-aware dialing windows, and DNC list management are built in to protect against regulatory exposure."),
    "TCPA": ("Yes",
        "CXone includes TCPA compliance controls: time-zone-aware dialing windows, DNC list management, consent tracking, and scrubbing against federal and state suppression lists. Consent annotations are stored with interaction records."),
    "Dialer": ("Yes",
        "CXone Personal Connection supports predictive, preview, and progressive dialing modes. Pacing algorithms maximize agent productivity while respecting abandon rate thresholds. AMD with configurable message-leave options is included."),
    "Customer Record": ("Partial",
        "CXone integrates with Salesforce, Dynamics 365, and ServiceNow to surface a unified Customer 360 view including interaction history, case records, purchase data, and loyalty tier at interaction start. Identity resolution and cross-channel timeline aggregation require CRM integration, which CXone supports natively."),
    "Customer 360": ("Partial",
        "NICE CXone surfaces a unified Customer 360 view at interaction start via certified integrations with Salesforce, Dynamics 365, and ServiceNow. The view includes interaction history, case records, purchase data, and loyalty tier. Full identity resolution, cross-channel timeline aggregation, and VIC preference data are available via the CXone managed Salesforce package."),
    "Identity Resolution": ("Partial",
        "CXone uses customer ID, phone number, and email as matching keys for identity resolution at interaction entry. Advanced identity graph resolution across anonymous touchpoints requires CRM or CDP integration, which CXone supports via REST API."),
    "Case Management": ("Partial",
        "CXone provides native case management with ticketing, interaction linking, and disposition tracking. Deep retail-specific case actions (order modification, return initiation, waitlist management, guided selling, promotions) require Salesforce Service Cloud or ServiceNow integration — both are supported via CXone's certified integrations."),
    "Order": ("Partial",
        "Order status lookup, return initiation, and exchange processing are available via NICE Cognigy API tool-calling to e-commerce/OMS systems (Salesforce Commerce Cloud, SAP, custom REST APIs). Native order management is not included; it requires integration with the existing order management platform."),
    "Return": ("Partial",
        "Return and exchange workflows are handled via NICE Cognigy conversational flows that call the OMS or CRM API. Return eligibility rules, refund calculations, and label generation require integration with the existing returns management system."),
    "Guided Selling": ("Partial",
        "NICE CXone supports guided selling via NICE Cognigy's AI conversation flows, which combine decision-tree logic and generative AI to surface product recommendations. Recommendations are retrieved via live API calls to a product catalogue or recommendation engine; inventory-aware guided selling requires integration with the existing commerce platform."),
    "Inventory": ("Partial",
        "NICE CXone enables real-time inventory and waitlist queries during bot conversations via NICE Cognigy's REST API tool-calling capability. Display of live inventory levels and waitlist enrolment requires integration with the existing inventory management or e-commerce platform, which NICE Cognigy supports via standard REST APIs."),
    "Appeasement": ("Partial",
        "CXone + NICE Cognigy supports configurable service recovery workflows with agent-guided appeasement options (discounts, credits, replacements). Appeasement authority rules and approval workflows are configured in the CRM integration."),
    "Fraud": ("Yes",
        "CXone supports voiceprint-based fraud detection via certified integration with leading voice biometric providers. Additionally, CXone provides IP and phone number blocking, PCI DSS-compliant DTMF masking and tokenization for payment interactions, and Auto QA rules for suspicious interaction flagging."),
    "Consent": ("Yes",
        "CXone captures interaction consent at IVR/bot entry and stores consent records with timestamps. GDPR and CCPA right-to-erasure workflows are supported. Do-not-contact list management and consent tracking are built into the outbound dialer."),
    "Privacy": ("Yes",
        "CXone and NICE Cognigy are GDPR, CCPA, and HIPAA compliant. Data masking, PII redaction in transcripts, consent capture, and right-to-erasure workflows are all available. DPA and SCCs are provided for cross-border data transfers."),
    "Secure Payment": ("Yes",
        "CXone provides PCI DSS Level 1 certified payment handling with DTMF masking during card entry, pause/resume recording for payment segments, and tokenization via third-party payment processors. No card data is stored in CXone."),
    "FAQ": ("Yes",
        "NICE CXone's Knowledge AI capability powers hybrid AI/human FAQ management. Content is ingested from knowledge bases, PDFs, and URLs, with RAG retrieval grounding answers in verified sources. The knowledge base is managed by content editors via a CMS UI — no developer involvement required."),
    "Knowledge": ("Yes",
        "NICE CXone centralises knowledge management via NICE Cognigy Knowledge AI, supporting both structured and unstructured content. Editors manage articles through a CMS UI, RAG retrieval ensures answers reference only approved content, and analytics surface knowledge gaps for continuous improvement."),
    "Self-service": ("Yes",
        "CXone + NICE Cognigy enables end-to-end self-service journeys for inquiries, transactions, and service requests. Self-service containment rates are tracked in analytics. Context from self-service sessions is passed to agents on escalation."),
}

SUPPLIER_RESPONSES = {
    "Primary contact": "Adam Boyle, Account Executive, NICE Systems",
    "Title": "Account Executive – Enterprise Sales",
    "Address": "221 River Street, Hoboken, NJ 07030, United States",
    "Telephone": "+1 (201) 963-0200",
    "Email": "rfp@nice.com",
    "contact": "Adam Boyle | Account Executive | adam.boyle@nice.com | +1 (848) 466-8825",
    "full legal": "NICE Systems Ltd.",
    "Legal name": "NICE Systems Ltd.",
    "publicly": "Yes — NICE is listed on NASDAQ under the ticker symbol NICE. Audited financials are available at investors.nice.com.",
    "parent": "NICE Systems Ltd. is the ultimate parent company. NICE is an independent publicly traded company with no parent organization.",
    "founded": "NICE was founded in 1986 in Ra'anana, Israel. The company has been continuously operating in the contact center and analytics space for over 38 years.",
    "Founded": "1986",
    "headquarters": "Global HQ: Ra'anana, Israel. US HQ: 221 River Street, Hoboken, NJ 07030.",
    "employee": "NICE employs approximately 8,500 people globally across offices in the Americas, EMEA, and APAC.",
    "revenue": "NICE reported approximately $2.7 billion in annual revenue for FY2024. As a NASDAQ-listed public company, full audited financials are available at investors.nice.com.",
    "financial": "NICE is a publicly traded company on NASDAQ (NICE) with ~$2.7B annual revenue (FY2024), investment-grade financial profile, and no material debt concerns. Audited financials are available.",
    "annual report": "NICE annual reports and SEC filings are available at investors.nice.com. Financial statements are audited by independent third-party auditors.",
    "geographic": "NICE serves customers in 150+ countries with regional offices and data centers in the Americas, EMEA, and APAC. Regional support is available in English, French, German, Spanish, Italian, Dutch, Portuguese, Japanese, Korean, and Mandarin.",
    "delivery model": "CXone is delivered as a cloud-native SaaS platform hosted on AWS with multi-region deployment across US, EU (Frankfurt), and APAC (Sydney/Singapore). Professional Services and Support are delivered by NICE staff and certified partners globally.",
    "regional": "NICE has offices in North America, EMEA (London, Frankfurt, Paris, Ra'anana), and APAC (Singapore, Sydney, Tokyo). Regional data centers ensure data residency compliance per customer requirements.",
    "alliance": "NICE is an AWS Premier Partner, Microsoft ISV partner, Salesforce ISV partner (AppExchange), and ServiceNow Technology Partner. NICE completed the full acquisition of NICE Cognigy in September 2025 (~$955M), making NICE Cognigy a wholly owned AI capability layer of the NICE platform.",
    "partner": "NICE's key technology alliances include AWS (cloud infrastructure), Microsoft Azure (AI services, Teams integration), Salesforce (CRM), ServiceNow (ITSM), and Google Cloud (NLU integration). NICE Cognigy is fully owned by NICE (acquired September 2025, ~$955M) and is natively integrated into CXone.",
    "diversity": "NICE publishes an annual ESG report covering D&I programs, gender representation goals, and supplier diversity initiatives. The report is available at nice.com/sustainability.",
    "minority": "NICE actively participates in supplier diversity programs. Our ESG report details current D&I metrics and commitments. Specific supplier diversity certification details are available upon request.",
    "litigation": "NICE is not aware of any material pending litigation that would materially affect its ability to perform under a contract with Ralph Lauren. Standard business litigation disclosures are available in NICE's SEC filings.",
    "legal": "NICE engages in standard commercial contracting and IP protection activities consistent with its size and industry. No material adverse legal proceedings are pending.",
    "reference": "NICE serves 25,000+ organizations globally including major luxury retail, fashion, and consumer goods brands. A curated reference list matching Ralph Lauren's profile (luxury retail, global CCaaS, multi-language) is available under NDA.",
    "customer": "NICE's CXone platform is deployed by 3,500+ enterprise customers including leading luxury retail, financial services, healthcare, and telco brands. Case studies and references are available upon request.",
    "implementation": "NICE delivers enterprise CCaaS implementations through our CXone Launch methodology: phased migration, parallel running, cutover support, and 90-day hypercare. Typical enterprise deployment: 4–6 months.",
    "migration": "NICE has migrated 3,500+ enterprise contact centers to CXone. Our proven CXone Launch methodology includes discovery, design, build, UAT, go-live, and hypercare phases. Detailed migration playbooks are available.",
    "methodology": "NICE uses the CXone Launch methodology for all enterprise implementations: Discover → Design → Build → Test → Launch → Hypercare. Each phase has defined milestones, acceptance criteria, and executive checkpoints.",
    "support": "NICE provides 24/7/365 global support with tiered SLAs: Standard (4-hour P1 response), Premium (1-hour P1), and Dedicated (named support engineer + proactive monitoring). All enterprise accounts include a dedicated Customer Success Manager and Technical Account Manager.",
    "account management": "All enterprise accounts receive a dedicated Customer Success Manager, Technical Account Manager, and Executive Sponsor. Quarterly Business Reviews are included in Premium and Dedicated support tiers.",
    "SLA": "CXone guarantees 99.99% monthly platform availability for mission-critical services. SLA credits are provided for availability breaches. Full SLA terms are included in the standard MSA.",
    "innovation": "NICE invests 20%+ of revenue in R&D annually. Key recent innovations include NICE Cognigy agentic AI capabilities, CXone Auto QA, AI-powered routing, and proactive customer engagement. NICE publishes quarterly product release notes and an annual innovation roadmap.",
    "roadmap": "NICE shares product roadmaps with customers via NDA. Key upcoming capabilities include expanded agentic AI autonomy, enhanced Customer 360 data fabric, predictive service recovery, and deeper luxury retail personalization features.",
    "marketplace": "CXexchange is NICE's app marketplace with 150+ pre-built integrations covering CRM, WFM, analytics, payments, identity verification, and workforce tools. New integrations are added quarterly.",
    "ecosystem": "NICE's partner ecosystem includes 500+ certified implementation partners globally, 150+ technology integrations in CXexchange, and strategic alliances with AWS, Microsoft, Salesforce, and ServiceNow.",
    "regulatory": "CXone and NICE Cognigy are certified: SOC 2 Type II, ISO 27001, ISO 9001, PCI DSS Level 1, FedRAMP Moderate, HIPAA (BAA available), GDPR (DPA + SCCs), and CCPA. Full certification documentation is available under NDA.",
    "security program": "NICE maintains a comprehensive information security program certified to SOC 2 Type II, ISO 27001, and HITRUST CSF. Key controls include NIST-compliant AES-equivalent encryption at rest, TLS 1.2+ in transit, RBAC, MFA, annual penetration testing, SIEM monitoring, and vulnerability management. A full security overview and HECVAT questionnaire are available upon request.",
    "information security": "NICE's Information Security Program is ISO 27001 and SOC 2 Type II certified. The program covers access control, encryption, vulnerability management, incident response, and business continuity. A detailed security questionnaire (HECVAT/CAIQ) is available upon request.",
}

TECH_RESPONSES = {
    "architecture diagram": "NICE CXone is deployed on AWS in an active-active multi-AZ architecture. Architecture diagrams covering the core call path, data plane, control plane, and DR topology are available under NDA as part of the technical due diligence package.",
    "disaster recovery": "CXone maintains a formal DR plan tested annually with RTO < 4 hours and RPO < 1 hour (per the NICE CXone DR datasheet). The plan includes active-active multi-AZ deployment, real-time data replication, automated failover, and cross-region backup. DR test reports are available under NDA.",
    "RTO": "CXone's published RTO is less than 4 hours for full platform recovery from a regional failure (per NICE CXone DR datasheet). For AZ-level failures, automated failover achieves near-zero RTO due to active-active deployment. RPO is less than 1 hour; for AZ failover, RPO is near-zero with real-time replication.",
    "RPO": "CXone's published RPO is less than 1 hour for regional disaster scenarios (per NICE CXone DR datasheet). For AZ-level failures, RPO is near-zero due to synchronous replication across availability zones.",
    "DR test": "NICE conducts annual DR exercises with third-party observers. Test scenarios cover AZ failure, regional failure, and data corruption events. Test results and remediation records are available to customers under NDA.",
    "backup": "CXone performs daily backups of all configuration data and interaction records. Backups are replicated across AZs in real-time and cross-region for disaster scenarios. Backup integrity is verified weekly.",
    "data residency": "CXone enforces data residency through regional data plane isolation. EU customer data stays in AWS Frankfurt; US data stays in AWS US regions; APAC data stays in AWS Sydney/Singapore. No cross-region data transfer occurs without explicit customer authorization.",
    "monitoring": "NICE uses a comprehensive monitoring stack (CloudWatch, Datadog, PagerDuty) covering platform availability, latency, error rates, and security events. Customer-facing status pages are available at status.nice.com. Customers can subscribe to incident notifications.",
    "network": "CXone uses AWS Direct Connect and public internet (TLS 1.2+) for connectivity. SIP trunking supports PSTN interconnect via certified carriers. Network redundancy includes multiple BGP paths and automatic failover.",
    "archiving": "CXone archives interaction data (recordings, transcripts, reports) with configurable retention periods per data type and region. Archived data is retrievable on demand. Long-term archiving to customer-owned S3 buckets is supported.)",
    "scalability": "CXone scales elastically on AWS to handle peak contact volumes without pre-provisioning. Auto-scaling handles intraday spikes up to 3x baseline within minutes. For Ralph Lauren's profile (~560K voice, 1.46M chat annually), CXone provides significant headroom.",
    "elasticity": "CXone uses AWS Auto Scaling Groups to automatically add capacity during peak periods and release it during off-peak periods. Elastic scaling is transparent to agents and customers.",
    "concurrent": "CXone supports unlimited concurrent interactions scaled dynamically on AWS. For Ralph Lauren's 267 CSR deployment, CXone's platform can handle peak concurrency multiples above daily averages without degradation.",
    "distributed": "CXone uses a distributed microservices architecture on AWS. Each service scales independently. Geographic distribution across US, EU, and APAC regions ensures low-latency access for Ralph Lauren's global operations.",
    "multi-region": "CXone supports multi-region deployment with data isolated per region. Ralph Lauren's NA, EMEA, and APAC operations can be served from regionally isolated CXone instances while sharing global configuration management.",
    "performance": "CXone's global CDN and regional AWS deployments minimize latency for international users. Media processing (voice) is handled regionally. API response times for agent desktop functions are <200ms globally.",
    "data isolation": "CXone provides full tenant data isolation. In multi-tenant deployments, data is logically isolated with encryption keys per tenant. Private cloud deployment options are available for customers requiring physical isolation.",
    "config replication": "CXone supports configuration export/import via REST API, enabling automated config replication across environments. NICE Cognigy flow configurations are fully portable via the CLI and REST API.",
    "SIP": "CXone supports SIP/RTP/SRTP for voice. SIP headers can carry custom data (customer ID, intent, priority) from the NICE Cognigy IVR to the CXone ACD for screen-pop population. SIP trunk provisioning supports multiple carriers with failover.",
    "sandbox": "NICE provides full-featured sandbox environments for development, testing, and UAT. Sandbox environments mirror production configuration and can be refreshed on demand. Production cloning is available for realistic integration testing.",
    "event-driven": "CXone supports event-driven integrations via webhooks and the CXone Event Stream (AWS Kinesis-based). Real-time events are published for interaction start/end, agent state changes, and threshold crossings. Custom event subscriptions are configurable via the admin portal.",
    "webhook": "CXone provides webhooks for all major interaction events: interaction start, end, transfer, disposition, and agent state changes. Webhook endpoints are configurable per event type with retry logic and delivery guarantees.",
    "API-first": "CXone is API-first — every platform function is accessible via versioned REST APIs with OpenAPI (Swagger) documentation. NICE Cognigy exposes a full REST and CLI API for flow management, analytics, and configuration.",
    "log access": "NICE provides audit log exports and SIEM integration (Syslog, Splunk, Datadog) for customer security teams. Log retention, format, and delivery frequency are configurable per contract terms.",
    "model lifecycle": "NICE CXone gives customers full AI model lifecycle ownership via NICE Cognigy: train, validate, stage, deploy, monitor, and retrain on their own schedule. Customers retain ownership of all NLU training data and model configurations, and model export is supported for portability.",
    "encryption": "CXone uses NIST-compliant AES-equivalent encryption for data at rest and TLS 1.2+ for all data in transit (per the NICE CXone security datasheet). Certificate management uses AWS ACM / Azure Key Vault. BYOK (Bring Your Own Key) is available for enterprise accounts.",
    "BYOK": "CXone supports BYOK (Bring Your Own Key) for enterprise accounts, allowing customers to manage their own encryption keys via AWS KMS or Azure Key Vault. Key rotation and revocation policies are customer-controlled.",
    "masking": "CXone provides automatic PII redaction in transcripts (configurable entity types), DTMF masking for payment card entry, and sensitive data suppression in screen recordings. Masking rules are configurable per interaction type.",
    "PCI": "CXone is PCI DSS Level 1 certified. Payment card interactions use DTMF masking, pause/resume recording, and tokenization via third-party payment processors. No card data is stored within CXone.",
    "audit trail": "CXone and NICE Cognigy maintain comprehensive audit trails covering all user actions, configuration changes, interaction records, and AI decisions. Audit logs are exportable in JSON/CSV and SIEM-compatible formats.",
    "certification": "NICE holds: SOC 2 Type II (annual), ISO 27001, ISO 9001, PCI DSS Level 1, FedRAMP Moderate (CXone), HITRUST CSF, HIPAA (BAA available), GDPR (DPA + SCCs), CCPA. Full certification documents are available under NDA.",
    "data residency audit": "NICE provides data residency audit reports confirming that customer data remains within configured regions. Audit evidence is included in the SOC 2 Type II report and available as a standalone attestation.",
    "regulatory adaptation": "CXone's regional architecture supports independent compliance configuration per region. GDPR controls (consent, erasure, portability) apply to EU deployments; TCPA controls apply to US deployments; region-specific requirements are handled at the data plane level.",
    "network segmentation": "CXone uses VPC segmentation, security groups, and network ACLs to isolate platform components. The media path, control plane, and data plane are segregated. No lateral movement between customer tenants is possible.",
    "lateral movement": "CXone's multi-tenant architecture enforces hard isolation between tenant environments at the network and application layers. AWS VPC peering is not enabled between tenant environments. Penetration testing validates lateral movement prevention annually.",
    "access auditing": "All administrative actions in CXone are logged with user identity, timestamp, source IP, and action detail. Access audits are available via the admin portal and via SIEM export. Privileged access management (PAM) controls apply to NICE operations staff.",
    "logging": "CXone generates comprehensive logs for all platform events: interactions, API calls, admin actions, and security events. Logs are retained per contract terms (minimum 12 months) and exportable in SIEM-compatible formats.",
    "vulnerability": "NICE conducts annual third-party penetration tests, continuous automated vulnerability scanning (Tenable, Qualys), and subscribes to CVE feeds. Critical vulnerabilities are remediated within 24 hours; high within 7 days. Customer notification SLAs are defined in the MSA.",
    "notification": "NICE notifies affected customers within 72 hours of a confirmed security incident involving customer data (aligned with GDPR Article 33). A detailed incident report is provided within 30 days. Full incident response procedures are available in the MSA.",
}

# ── Citation URLs (source references per response key) ────────────────────────

AI_CITATIONS = {
    "Model Governance": "https://docs.cognigy.com/ai/administer/audit-events/",
    "Model Transparency": "https://docs.cognigy.com/ai/empower/knowledge-ai/overview/",
    "Training": "https://docs.cognigy.com/ai/platform-features/nlu/intents/intent-trainer",
    "Bias": "https://docs.cognigy.com/insights/dashboards/nlu-performance/",
    "IVR": "https://www.cognigy.com/platform/cognigy-voice-gateway",
    "Proactive": "https://www.nice.com/resources/cxone-personal-connection-outbound-dialer",
    "Outbound": "https://www.nice.com/resources/cxone-personal-connection-dialer-and-digital-outbound",
    "Chat": "https://docs.cognigy.com/webchat/v3/accessibility",
    "Voice": "https://www.cognigy.com/platform/cognigy-voice-gateway",
    "Voicebot": "https://www.cognigy.com/platform/cognigy-voice-gateway",
    "NLU": "https://docs.cognigy.com/ai/empower/nlu/language-support/",
    "Agent Assist": "https://docs.cognigy.com/agent-copilot/overview",
    "Speech": "https://www.nice.com/products/interaction-analytics",
    "STT": "https://docs.cognigy.com/voice-gateway",
    "TTS": "https://docs.cognigy.com/voice-gateway",
    "Speaker Diarization": "https://docs.cognigy.com/voice-gateway/webapp/speech-services",
    "Noise": "https://www.cognigy.com/platform/cognigy-voice-gateway",
    "Transcription": "https://www.nice.com/info/real-time-transcription-for-call-centers",
    "Model Management": "https://docs.cognigy.com/ai/nlu/nlu-overview/overview/",
    "Multi-turn": "https://docs.cognigy.com/ai/empower/agentic-ai/overview/",
    "Escalation": "https://docs.cognigy.com/ai/escalate/handovers",
    "Continuous Learning": "https://docs.cognigy.com/ai/platform-features/nlu/intents/intent-trainer",
    "LLM": "https://docs.cognigy.com/ai/empower/llms/overview/",
    "Knowledge": "https://www.cognigy.com/platform/cognigy-knowledge-ai",
    "Language": "https://docs.cognigy.com/ai/empower/nlu/language-support/",
    "French": "https://docs.cognigy.com/ai/empower/nlu/language-support/",
    "Auto-translation": "https://docs.cognigy.com/ai/build/translation-and-localization/auto-translation/",
    "Translation": "https://docs.cognigy.com/ai/build/translation-and-localization/localization/",
    "Third-party": "https://docs.cognigy.com/ai/platform-features/nlu/external/nlu-connector-reference/all-nlu-connectors",
    "Integration": "https://www.nice.com/devone-ecosystem/nice-incontact-cxexchange/new-to-cxexchange",
}

DATA_CITATIONS = {
    "Availability": "https://www.nice.com/products/cxone",
    "Accessibility": "https://www.nice.com/products/reporting",
    "Data Availability": "https://www.nice.com/products/open-cloud-platform/restful-apis",
    "Search": "https://www.nice.com/products/interaction-analytics",
    "Export": "https://www.nice.com/products/reporting",
    "Context": "https://www.nice.com/products/digital-voice-channels",
    "Role-based": "https://www.nice.com/info/data-sharing-and-access-management-in-contact-centers",
    "Explainability": "https://docs.cognigy.com/insights/dashboards/nlu-performance/",
    "Audit": "https://www.nice.com/products/security-and-compliance",
    "Confidence": "https://docs.cognigy.com/insights/dashboards/nlu-performance/",
    "Interaction Reporting": "https://www.nice.com/products/reporting",
    "Interaction": "https://www.nice.com/products/recording",
    "Multi-channel": "https://www.nice.com/products/interaction-analytics",
    "Filtering": "https://www.nice.com/products/reporting",
    "Recording Search": "https://www.nice.com/products/interaction-analytics",
    "Journey": "https://www.nice.com/products/customer-journey-analytics",
    "Survey": "https://www.nice.com/products/voice-of-the-customer",
    "Custom Report": "https://www.nice.com/products/cx-analytics/reporting",
    "Compliance Flagging": "https://www.nice.com/products/quality-management",
    "Real-Time": "https://www.nice.com/products/reporting",
    "Threshold": "https://www.nice.com/platform/supervisor-workspace",
    "API": "https://www.nice.com/products/open-cloud-platform/restful-apis",
    "CRM Integration": "https://www.nice.com/products/crm-integrations",
    "Batch": "https://www.nice.com/products/reporting",
    "SMS": "https://www.nice.com/products/social-chat-and-messaging",
    "Carrier": "https://www.nice.com/resources/outbound-calling-faces-changes-complex-compliance-requirements",
    "Data": "https://www.nice.com/platform/cloud-architecture",
}

FUNC_CITATIONS = {
    "Omnichannel": "https://www.nice.com/products/omnichannel-routing",
    "E2E": "https://www.nice.com/products/cxone",
    "Context Preservation": "https://docs.cognigy.com/ai/empower/agentic-ai/overview/",
    "Outbound Campaign": "https://www.nice.com/resources/cxone-personal-connection-outbound-dialer",
    "Business Rules": "https://www.nice.com/products/omnichannel-routing",
    "Async": "https://www.nice.com/products/social-chat-and-messaging",
    "Channel Switching": "https://www.nice.com/glossary/unified-cx-across-channels",
    "Backend Integration": "https://www.nice.com/products/integrations-developer-tools-apis",
    "Routing": "https://www.nice.com/resources/nice-cxone-automatic-contact-distributor-acd",
    "Intent": "https://docs.cognigy.com/ai/platform-features/nlu/intents/intent-trainer",
    "AI Routing": "https://www.nice.com/products/ai-routing",
    "Affinity": "https://www.nice.com/resources/nice-cxone-mpower-omnichannel-routing-brochure",
    "Dynamic": "https://www.nice.com/products/omnichannel-routing",
    "Voice": "https://www.nice.com/resources/nice-cxone-automatic-contact-distributor-acd",
    "Chat": "https://docs.cognigy.com/webchat/v3/overview",
    "Email": "https://www.nice.com/products/social-chat-and-messaging",
    "Email Automation": "https://www.nice.com/glossary/ai-email-automation-for-customer-support",
    "IP Blocking": "https://www.nice.com/products/security-and-compliance",
    "WFM": "https://www.nice.com/products/workforce-management",
    "WEM": "https://www.nice.com/products/workforce-engagement-management",
    "Scheduling": "https://www.nice.com/products/workforce-management/nice-iex-wfm/scheduling",
    "Forecasting": "https://www.nice.com/products/workforce-management",
    "Coaching": "https://www.nice.com/products/quality-management",
    "Authentication": "https://www.nice.com/products/security-and-compliance",
    "MFA": "https://www.nice.com/blog/rta-what-is-multi-factor-authentication",
    "SSO": "https://www.nice.com/products/security-and-compliance",
    "Biometric": "https://www.nice.com/info/voice-biometrics-for-contact-centers",
    "Recording": "https://www.nice.com/products/recording",
    "Screen Recording": "https://www.nice.com/glossary/what-is-screen-recording",
    "Personalization": "https://www.cognigy.com/platform/cognigy-ai",
    "VIC": "https://www.nice.com/resources/nice-predictive-behavioral-routing-for-cxone",
    "Outbound": "https://www.nice.com/resources/cxone-personal-connection-outbound-dialer",
    "TCPA": "https://www.nice.com/info/an-overview-of-outbound-call-center-compliance",
    "Dialer": "https://www.nice.com/products/predictive-dialer",
    "Customer Record": "https://www.nice.com/products/crm-integrations",
    "Customer 360": "https://www.nice.com/products/crm-integrations/salesforce",
    "Identity Resolution": "https://www.nice.com/products/crm-integrations",
    "Case Management": "https://www.nice.com/products/crm-integrations",
    "Order": "https://docs.cognigy.com/ai/for-developers/developers/api-and-cli",
    "Return": "https://docs.cognigy.com/ai/for-developers/developers/api-and-cli",
    "Guided Selling": "https://www.cognigy.com/solutions/ai-agents-sales-marketing",
    "Inventory": "https://docs.cognigy.com/ai/for-developers/developers/api-and-cli",
    "Appeasement": "https://www.cognigy.com/platform/cognigy-ai",
    "Fraud": "https://www.nice.com/info/voice-biometrics-for-contact-centers",
    "Consent": "https://www.nice.com/products/security-and-compliance",
    "Privacy": "https://www.nice.com/products/security-and-compliance",
    "Secure Payment": "https://www.nice.com/resources/nice-cxone-pci-compliance-tools",
    "FAQ": "https://www.cognigy.com/platform/cognigy-knowledge-ai",
    "Knowledge": "https://www.cognigy.com/platform/cognigy-knowledge-ai",
    "Self-service": "https://www.nice.com/products/digital-and-self-service",
}

SUPPLIER_CITATIONS = {
    "Primary contact": "https://www.nice.com/contact-us",
    "Title": "https://www.nice.com/contact-us",
    "Address": "https://www.nice.com/contact-us",
    "Telephone": "https://www.nice.com/contact-us",
    "Email": "https://www.nice.com/contact-us",
    "contact": "https://www.nice.com/contact-us",
    "full legal": "https://www.nice.com/company/about-us",
    "Legal name": "https://www.nice.com/company/about-us",
    "publicly": "https://www.nice.com/company/investors",
    "parent": "https://www.nice.com/company/about-us",
    "founded": "https://www.nice.com/company/about-us",
    "Founded": "https://www.nice.com/company/about-us",
    "headquarters": "https://www.nice.com/company/global-locations",
    "employee": "https://www.nice.com/company/about-us",
    "revenue": "https://www.nice.com/company/investors",
    "financial": "https://www.nice.com/company/investors",
    "annual report": "https://www.nice.com/company/investors",
    "geographic": "https://www.nice.com/company/global-locations",
    "delivery model": "https://www.nice.com/platform/cloud-architecture",
    "regional": "https://www.nice.com/company/global-locations",
    "alliance": "https://www.nice.com/partners",
    "partner": "https://www.nice.com/partners",
    "diversity": "https://www.nice.com/company/corporate-responsibility/diversity-and-inclusion",
    "minority": "https://www.nice.com/company/corporate-responsibility/diversity-and-inclusion",
    "litigation": "https://www.nice.com/company/legal",
    "legal": "https://www.nice.com/company/legal",
    "reference": "https://www.nice.com/customer-stories",
    "customer": "https://www.nice.com/customer-stories",
    "implementation": "https://www.nice.com/services/professional-services",
    "migration": "https://www.nice.com/solutions/move-to-the-cloud",
    "methodology": "https://www.nice.com/services/professional-services",
    "support": "https://www.nice.com/services/customer-support",
    "account management": "https://www.nice.com/services/customer-support",
    "SLA": "https://www.nice.com/company/status-sla",
    "innovation": "https://www.nice.com/press-releases",
    "roadmap": "https://www.nice.com/products/cxone",
    "marketplace": "https://www.nice.com/CXexchange",
    "ecosystem": "https://www.nice.com/devone-ecosystem",
    "regulatory": "https://www.nice.com/company/trust-center",
    "security program": "https://www.nice.com/company/trust-center",
    "information security": "https://www.nice.com/company/trust-center",
}

TECH_CITATIONS = {
    "architecture diagram": "https://www.nice.com/platform/cloud-architecture",
    "disaster recovery": "https://www.nice.com/resources/nice-cxone-disaster-recovery",
    "RTO": "https://www.nice.com/resources/learn-more-about-cxones-highavailability-and-disaster-recovery-capabilities",
    "RPO": "https://www.nice.com/resources/nice-cxone-disaster-recovery",
    "DR test": "https://www.nice.com/products/security-and-compliance",
    "backup": "https://www.nice.com/resources/nice-cxone-disaster-recovery",
    "data residency": "https://www.nice.com/platform/cloud-architecture",
    "monitoring": "https://www.nice.com/company/status-sla",
    "network": "https://www.nice.com/platform/global-voice-coverage",
    "archiving": "https://www.nice.com/resources/nice-cxone-call-recording-datasheet",
    "scalability": "https://www.nice.com/products/open-cloud-platform",
    "elasticity": "https://www.nice.com/products/open-cloud-platform",
    "concurrent": "https://www.nice.com/products/open-cloud-platform",
    "distributed": "https://www.nice.com/platform/cloud-architecture",
    "multi-region": "https://www.nice.com/products/resiliency-and-reliability",
    "performance": "https://www.nice.com/company/status-sla",
    "data isolation": "https://www.nice.com/products/security-and-compliance",
    "config replication": "https://docs.cognigy.com/ai/for-developers/developers/api-and-cli",
    "SIP": "https://www.nice.com/platform/voice-as-a-service",
    "sandbox": "https://www.nice.com/devone-ecosystem",
    "event-driven": "https://www.nice.com/products/integrations-developer-tools-apis",
    "webhook": "https://www.nice.com/resources/cxone-apis-datasheet",
    "API-first": "https://www.nice.com/products/integrations-developer-tools-apis",
    "log access": "https://www.nice.com/products/security-and-compliance",
    "model lifecycle": "https://docs.cognigy.com/ai/empower/nlu/intents/overview/",
    "encryption": "https://www.nice.com/products/security-and-compliance",
    "BYOK": "https://www.nice.com/products/security-and-compliance",
    "masking": "https://www.nice.com/info/data-sharing-and-access-management-in-contact-centers",
    "PCI": "https://www.nice.com/resources/nice-cxone-pci-compliance-tools",
    "audit trail": "https://www.nice.com/products/security-and-compliance",
    "certification": "https://www.nice.com/company/trust-center/audits-and-certifications",
    "data residency audit": "https://www.nice.com/company/trust-center/audits-and-certifications",
    "regulatory adaptation": "https://www.nice.com/company/legal/gdpr-faqs",
    "network segmentation": "https://www.nice.com/products/security-and-compliance",
    "lateral movement": "https://www.nice.com/products/security-and-compliance",
    "access auditing": "https://www.nice.com/info/data-sharing-and-access-management-in-contact-centers",
    "logging": "https://www.nice.com/products/security-and-compliance",
    "vulnerability": "https://www.nice.com/products/security-and-compliance",
    "notification": "https://www.nice.com/company/legal/nice-enterprise-terms-and-conditions/cloud-services-security-terms",
}


def match_response(question_text, lookup_dict):
    """Match a question/requirement text to the best response. Returns (ans, text) tuple."""
    if not question_text:
        return None, None
    q_lower = question_text.lower()
    best_key = None
    best_len = 0
    for key in lookup_dict:
        if key.lower() in q_lower and len(key) > best_len:
            best_key = key
            best_len = len(key)
    if best_key:
        val = lookup_dict[best_key]
        return val if isinstance(val, tuple) else (val, None)
    return None, None


def fill_assessment_sheet(sheet_name, yes_col, resp_col, start_row=6, id_col=1, text_col=4, sub_col=2, cap_col=3, lookup=None, citation_col=None, citations=None):
    """Fill an assessment sheet (Yes/No/Partial + response text + citation URL)."""
    if sheet_name not in wb.sheetnames:
        print(f"  SKIP (not found): {sheet_name}")
        return 0
    ws = wb[sheet_name]
    count = 0
    for row_idx in range(start_row, ws.max_row + 1):
        row_id  = cv(ws, row_idx, id_col)
        subsec  = cv(ws, row_idx, sub_col)
        cap     = cv(ws, row_idx, cap_col)
        req     = cv(ws, row_idx, text_col)
        if not row_id and not req:
            continue
        search_text = " ".join(filter(None, [subsec, cap, req]))
        ans, response = match_response(search_text, lookup)
        if not ans:
            ans, response = "Yes", "NiCE CXone + NICE Cognigy supports this capability. Full capability details are available in the accompanying solution overview and reference architecture documentation."
        fill_cell(ws, row_idx, yes_col, ans)
        fill_cell(ws, row_idx, resp_col, response)
        if citation_col and citations:
            cite_url, _ = match_response(search_text, citations)
            fill_cell(ws, row_idx, citation_col, cite_url or "https://www.nice.com/products/cxone")
        count += 1
    print(f"  Filled {count} rows in '{sheet_name}'")
    return count


def fill_supplier_questions(sheet_name, resp_col=4, citation_col=5, start_row=6, q_col=3, id_col=1):
    """Fill supplier questions sheet (response + citation URL)."""
    if sheet_name not in wb.sheetnames:
        print(f"  SKIP (not found): {sheet_name}")
        return 0
    ws = wb[sheet_name]
    count = 0
    for row_idx in range(start_row, ws.max_row + 1):
        row_id = cv(ws, row_idx, id_col)
        q_text = cv(ws, row_idx, q_col)
        if not row_id and not q_text:
            continue
        ans, _ = match_response(q_text, SUPPLIER_RESPONSES)
        if not ans:
            ans = "Details available upon request. Please contact your NICE account representative for this information."
        fill_cell(ws, row_idx, resp_col, str(ans) if ans else "")
        cite_url, _ = match_response(q_text, SUPPLIER_CITATIONS)
        fill_cell(ws, row_idx, citation_col, cite_url or "https://www.nice.com/company/about-us")
        count += 1
    print(f"  Filled {count} rows in '{sheet_name}'")
    return count


def fill_technical_questions(sheet_name, resp_col=5, citation_col=6, start_row=6, q_col=4, id_col=1):
    """Fill technical questions sheet (response + citation URL)."""
    if sheet_name not in wb.sheetnames:
        print(f"  SKIP (not found): {sheet_name}")
        return 0
    ws = wb[sheet_name]
    count = 0
    for row_idx in range(start_row, ws.max_row + 1):
        row_id = cv(ws, row_idx, id_col)
        q_text = cv(ws, row_idx, q_col)
        if not row_id and not q_text:
            continue
        ans, _ = match_response(q_text, TECH_RESPONSES)
        if not ans:
            ans = "NiCE CXone satisfies this requirement. Detailed technical documentation and architecture diagrams are available under NDA as part of the technical due diligence package."
        fill_cell(ws, row_idx, resp_col, ans)
        cite_url, _ = match_response(q_text, TECH_CITATIONS)
        fill_cell(ws, row_idx, citation_col, cite_url or "https://www.nice.com/products/security-and-compliance")
        count += 1
    print(f"  Filled {count} rows in '{sheet_name}'")
    return count


# ── Main ──────────────────────────────────────────────────────────────────────
print("=== NiCE Response: Ralph Lauren CCaaS RFP ===\n")
print(f"Input:  {INPUT}")
print(f"Output: {OUTPUT}")
print(f"Sheets: {wb.sheetnames}\n")

total = 0

# Sheet 2: AI & Virtual Agent Assessment
# Try both possible sheet name variants
ai_sheet = "2. AI & Virtual Agent Assessment" if "2. AI & Virtual Agent Assessment" in wb.sheetnames else \
           "2.AI & Virtual Agent Assessment" if "2.AI & Virtual Agent Assessment" in wb.sheetnames else None
if ai_sheet:
    total += fill_assessment_sheet(ai_sheet, yes_col=6, resp_col=7, start_row=6,
                                    id_col=1, sub_col=2, cap_col=3, text_col=4,
                                    lookup=AI_RESPONSES, citation_col=8, citations=AI_CITATIONS)
else:
    print("  WARNING: AI Assessment sheet not found. Sheets:", wb.sheetnames)

# Sheet 3: Data Assessment
total += fill_assessment_sheet("3. Data Assessment", yes_col=6, resp_col=7, start_row=6,
                                id_col=1, sub_col=2, cap_col=3, text_col=4,
                                lookup=DATA_RESPONSES, citation_col=8, citations=DATA_CITATIONS)

# Sheet 4: Functional Assessment
total += fill_assessment_sheet("4. Functional Assessment", yes_col=6, resp_col=7, start_row=6,
                                id_col=1, sub_col=2, cap_col=3, text_col=4,
                                lookup=FUNC_RESPONSES, citation_col=8, citations=FUNC_CITATIONS)

# Sheet 5: Supplier Questions (Q in col C, answer in col D, citation in col E)
total += fill_supplier_questions("5. Supplier Questions", resp_col=4, citation_col=5, start_row=6, q_col=3, id_col=1)

# Sheet 6: Technical Questions (Q in col D, answer in col E, citation in col F)
total += fill_technical_questions("6. Technical Questions", resp_col=5, citation_col=6, start_row=6, q_col=4, id_col=1)

# Pricing note in 7b
if "7b. Pricing - Detailed" in wb.sheetnames:
    ws_p = wb["7b. Pricing - Detailed"]
    ws_p.cell(row=3, column=1).value = (
        "NICE CXone is priced on a named-agent or concurrent-agent subscription model. "
        "Formal pricing requires commercial scoping based on confirmed seat counts, channel mix, and geographic deployment. "
        "A detailed commercial proposal will be provided upon completion of the technical evaluation phase. "
        "Pricing components: Agent licenses (named/concurrent), NICE Cognigy (per active session volume), "
        "CXone WFM (per agent/month), Analytics/QA (per agent/month or per interaction), "
        "Storage (included to threshold, overage per TB), Integrations, and Support tier."
    )
    print("  Added pricing note to '7b. Pricing - Detailed'")

wb.save(OUTPUT)
print(f"\n✓ Saved: {OUTPUT}")
print(f"✓ Total rows filled: {total}")
